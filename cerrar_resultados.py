"""
cerrar_resultados.py
---------------------
FASE 4, versión 4. Reintenta cada 15 min entre las 23:00 y la 01:00.

Hace 4 cosas (se agregó la 4ta en esta versión):
  1. Consulta el resultado final de cada partido seleccionado hoy y
     calcula si acertó el favorito (ganó) o no.
  2. NUEVO -- actualiza el rating propio (Glicko-2) de ambos equipos con
     el resultado real observado. Esto es lo que hace que el rating
     propio "aprenda" partido a partido, ganando confianza (RD bajando)
     con el tiempo -- el corazón del seguimiento continuo que se pidió.
  3. NUEVO -- audita el acierto de CADA ALERTA INDIVIDUAL enviada
     durante el partido, no solo el resultado final del favorito. Para
     cada alerta, revisa si hubo un gol real (del lado que la alerta
     predijo, o de cualquiera si era "partido_abierto") dentro de los
     15 minutos siguientes al minuto de la alerta -- usando el marcador
     guardado en cada snapshot. Sin esto, es imposible calibrar
     UMBRAL_GOL_INMINENTE u otros parámetros con evidencia real.
  4. Archiva el día completo y actualiza data/estadisticas.xlsx (ahora
     con una pestaña adicional de acierto por tipo de alerta).
"""

import json
from pathlib import Path

from fetch_data import obtener_resultado_fixture
from cuota_api_football import uso_de_hoy
from estado_diario import ya_se_hizo, marcar_hecho
import ratings_store

DATA_DIR = Path(__file__).parent / "data"
ARCHIVO_PARTIDOS = DATA_DIR / "partidos_hoy.json"
DIR_HISTORIAL_DIAS = DATA_DIR / "historial_dias"
ARCHIVO_EXCEL = DATA_DIR / "estadisticas.xlsx"

ESTADOS_TERMINADO = ("FT", "AET", "PEN")
VENTANA_ACIERTO_MINUTOS = 15


def _calcular_acierto(p, goles_local, goles_visitante):
    goles_favorito = goles_local if p["favorito_es_local"] else goles_visitante
    goles_rival = goles_visitante if p["favorito_es_local"] else goles_local
    return goles_favorito > goles_rival


def _actualizar_rating_propio(p, goles_local, goles_visitante):
    """Alimenta el rating propio (Glicko-2) de ambos equipos con el
    resultado real -- el paso que hace que el seguimiento sea continuo,
    no solo un snapshot inicial del día."""
    if goles_local > goles_visitante:
        resultado_local, resultado_visitante = 1.0, 0.0
    elif goles_local < goles_visitante:
        resultado_local, resultado_visitante = 0.0, 1.0
    else:
        resultado_local, resultado_visitante = 0.5, 0.5

    llave_local = ratings_store.llave_equipo(p.get("home_id"), nombre=p["local"])
    llave_visitante = ratings_store.llave_equipo(p.get("away_id"), nombre=p["visitante"])

    eq_local = ratings_store.obtener_o_crear(llave_local, nombre=p["local"])
    eq_visitante = ratings_store.obtener_o_crear(llave_visitante, nombre=p["visitante"])

    rating_local_antes, rd_local_antes = eq_local["rating"], eq_local["rd"]
    rating_visitante_antes, rd_visitante_antes = eq_visitante["rating"], eq_visitante["rd"]

    ratings_store.actualizar_tras_partido(llave_local, rating_visitante_antes, rd_visitante_antes,
                                           resultado_local, es_bootstrap=False)
    ratings_store.actualizar_tras_partido(llave_visitante, rating_local_antes, rd_local_antes,
                                           resultado_visitante, es_bootstrap=False)


def _hubo_gol_en_ventana(historial_snapshots, minuto_alerta, lado, favorito_es_local):
    """
    Revisa el marcador guardado en los snapshots para ver si el lado
    indicado ('favorito', 'rival' o 'cualquiera') anotó dentro de los
    VENTANA_ACIERTO_MINUTOS minutos siguientes al minuto de la alerta.
    Esta es la pieza que faltaba para calibrar el sistema con evidencia
    real en vez de solo el resultado final del partido.
    """
    snap_en_alerta = None
    snap_limite = None
    for s in historial_snapshots:
        if s["minuto"] <= minuto_alerta:
            snap_en_alerta = s
        if s["minuto"] <= minuto_alerta + VENTANA_ACIERTO_MINUTOS:
            snap_limite = s

    if snap_en_alerta is None or snap_limite is None:
        return None  # no hay suficientes datos para evaluar esta alerta

    gl_antes, gv_antes = snap_en_alerta["goles_local"], snap_en_alerta["goles_visitante"]
    gl_despues, gv_despues = snap_limite["goles_local"], snap_limite["goles_visitante"]

    gol_local = gl_despues > gl_antes
    gol_visitante = gv_despues > gv_antes

    if lado == "cualquiera":
        return gol_local or gol_visitante

    gol_favorito = gol_local if favorito_es_local else gol_visitante
    gol_rival = gol_visitante if favorito_es_local else gol_local

    if lado == "favorito":
        return gol_favorito
    if lado == "rival":
        return gol_rival
    return None


LADO_ESPERADO_POR_TIPO = {
    "posible_empate": "favorito",
    "posible_victoria_favorito": "favorito",
    "gana_favorito_1er_tiempo": "favorito",
    "ampliacion_marcador": "favorito",
    "posible_gol_no_favorito": "rival",
    "cuidado_rival_presiona": "rival",
    "gol_de_cierre": "favorito",
    "partido_abierto": "cualquiera",
    "tarjeta_roja": None,   # informativo, no predice un gol específico
    "penal": "cualquiera",  # el penal favorece fuertemente un gol, de cualquiera de los dos lados posibles según quién lo tira
    "partido_resuelto": None,
}


def _auditar_alertas(p):
    """Agrega 'acierto' (True/False/None) a cada alerta en
    p['alertas_enviadas'], comparando la predicción contra el marcador
    real observado en los minutos siguientes."""
    for alerta in p.get("alertas_enviadas", []):
        if "acierto" in alerta:
            continue  # ya auditada en un cierre anterior (no debería pasar, pero por si acaso)
        lado = LADO_ESPERADO_POR_TIPO.get(alerta["tipo"])
        if lado is None:
            alerta["acierto"] = None
            continue
        alerta["acierto"] = _hubo_gol_en_ventana(
            p["historial_snapshots"], alerta["minuto"], lado, p["favorito_es_local"])


def cerrar():
    if ya_se_hizo("cierre"):
        print("El cierre de hoy ya se hizo antes. Nada que hacer.")
        return

    if not ARCHIVO_PARTIDOS.exists():
        print("No hay partidos_hoy.json todavía. Se reintentará en el próximo ciclo.")
        return

    datos = json.loads(ARCHIVO_PARTIDOS.read_text(encoding="utf-8"))
    cambios = False

    for p in datos["partidos"]:
        if p.get("acierto") is not None or not p.get("fixture_id"):
            continue
        try:
            info = obtener_resultado_fixture(p["fixture_id"])
        except Exception as e:
            print(f"[AVISO] No se pudo consultar el resultado de {p['partido']}: {e}")
            continue
        if not info:
            continue

        estado = info["fixture"]["status"]["short"]
        if estado not in ESTADOS_TERMINADO:
            continue

        gh, ga = info["goals"]["home"], info["goals"]["away"]
        p["resultado_final"] = f"{gh}-{ga}"
        p["acierto"] = _calcular_acierto(p, gh, ga)

        _actualizar_rating_propio(p, gh, ga)
        _auditar_alertas(p)

        cambios = True

    if cambios:
        ARCHIVO_PARTIDOS.write_text(json.dumps(datos, ensure_ascii=False, indent=2), encoding="utf-8")

    usadas, disponibles = uso_de_hoy()

    DIR_HISTORIAL_DIAS.mkdir(exist_ok=True, parents=True)
    archivo_dia = DIR_HISTORIAL_DIAS / f"{datos['fecha']}.json"
    archivo_dia.write_text(json.dumps({
        "fecha": datos["fecha"],
        "partidos": datos["partidos"],
        "api_football_usadas": usadas,
        "api_football_disponibles": disponibles,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Día archivado en {archivo_dia}")

    _actualizar_excel(datos["fecha"], datos["partidos"], usadas, disponibles)
    marcar_hecho("cierre")


def _actualizar_excel(fecha, partidos, usadas, disponibles):
    try:
        import openpyxl
    except ImportError:
        print("[AVISO] openpyxl no está instalado, no se pudo actualizar el Excel "
              "(agrega 'openpyxl' a requirements.txt).")
        return

    if ARCHIVO_EXCEL.exists():
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        hoja1 = wb.create_sheet("Resultados diarios")
        hoja1.append(["Fecha", "Partido", "Favorito", "Local/Visitante", "Cuota inicial (proxy)",
                      "Prob. inicial %", "Marcador final", "Acierto", "Alertas enviadas",
                      "Rating propio local (n)", "Rating propio visitante (n)", "País verificado"])
        hoja2 = wb.create_sheet("Resumen por dia")
        hoja2.append(["Fecha", "Total partidos", "Aciertos", "% Aciertos",
                      "Peticiones API usadas", "Peticiones disponibles"])
        hoja3 = wb.create_sheet("Acierto por tipo de alerta")
        hoja3.append(["Fecha", "Tipo de alerta", "Enviadas", "Aciertos", "% Acierto"])

    hoja1 = wb["Resultados diarios"]
    hoja2 = wb["Resumen por dia"]
    hoja3 = wb["Acierto por tipo de alerta"] if "Acierto por tipo de alerta" in wb.sheetnames \
        else wb.create_sheet("Acierto por tipo de alerta")
    if hoja3.max_row == 1 and hoja3["A1"].value is None:
        hoja3.append(["Fecha", "Tipo de alerta", "Enviadas", "Aciertos", "% Acierto"])

    ya_registrado = any(fila[0].value == fecha for fila in hoja2.iter_rows(min_row=2) if fila[0].value)
    if ya_registrado:
        print(f"El día {fecha} ya estaba registrado en el Excel, no se duplica.")
        return

    total = len(partidos)
    aciertos = sum(1 for p in partidos if p.get("acierto") is True)
    resueltos = sum(1 for p in partidos if p.get("acierto") is not None)

    conteo_por_tipo = {}
    for p in partidos:
        hoja1.append([
            fecha, p["partido"], p["favorito"],
            "Local" if p["favorito_es_local"] else "Visitante",
            p["cuota_inicial"], p["probabilidad_inicial"],
            p.get("resultado_final") or "sin resolver",
            "✅" if p.get("acierto") is True else ("❌" if p.get("acierto") is False else "?"),
            len(p.get("alertas_enviadas", [])),
            p.get("rating_propio_partidos_local", 0), p.get("rating_propio_partidos_visitante", 0),
            "sí" if p.get("pais_verificado") else "no",
        ])
        for alerta in p.get("alertas_enviadas", []):
            tipo = alerta["tipo"]
            conteo_por_tipo.setdefault(tipo, {"enviadas": 0, "aciertos": 0})
            conteo_por_tipo[tipo]["enviadas"] += 1
            if alerta.get("acierto") is True:
                conteo_por_tipo[tipo]["aciertos"] += 1

    for tipo, c in conteo_por_tipo.items():
        pct = round((c["aciertos"] / c["enviadas"]) * 100, 1) if c["enviadas"] else None
        hoja3.append([fecha, tipo, c["enviadas"], c["aciertos"], pct])

    pct_aciertos = round((aciertos / resueltos) * 100, 1) if resueltos else None
    hoja2.append([fecha, total, aciertos, pct_aciertos, usadas, disponibles])

    DATA_DIR.mkdir(exist_ok=True)
    wb.save(ARCHIVO_EXCEL)
    print(f"Excel actualizado: {ARCHIVO_EXCEL} ({total} partidos, {aciertos} aciertos)")


if __name__ == "__main__":
    cerrar()
